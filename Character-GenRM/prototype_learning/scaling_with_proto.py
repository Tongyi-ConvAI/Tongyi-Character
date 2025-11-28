# -*- coding: utf-8 -*-
import json
import random
import csv
from pathlib import Path
import argparse
import concurrent.futures
import threading
import re  # ← Added: clean illegal chars
import time  # ← Added: slight jitter rate limiting

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from tqdm import tqdm

from pre_process.dataset import PrismDataset, ChatbotArenaDataset,Prism_personal_align_Dataset
from proto_knn_scaler import ProtoKNNUpliftScorer

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from inference.PROMPTS import USER_PREFERENCE_ANALYSIS_PROMPT,USER_PROMPT_TEMPLATE

from utils import fewshot_formatter, get_full_train_val_rotate

# Added: xlsx writing dependency
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError  # ← Added: exception type

# ---------------------------
# Thread-safe printing
# ---------------------------
_print_lock = threading.Lock()
def log(msg: str):
    with _print_lock:
        print(msg, flush=True)

# ---------------------------
# Robust HTTP Session (POST will also retry)
# ---------------------------
def make_robust_session(
    total=5,
    connect=5,
    read=5,
    status=5,
    backoff_factor=0.5,
    status_forcelist=(429, 500, 502, 503, 504),
    allowed_methods=frozenset(['GET', 'POST', 'PUT', 'DELETE', 'HEAD', 'OPTIONS', 'TRACE', 'PATCH']),
    pool_connections=200,
    pool_maxsize=200,
    base_headers=None,
):
    s = requests.Session()
    retries = Retry(
        total=total,
        connect=connect,
        read=read,
        status=status,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
        allowed_methods=allowed_methods,
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(pool_connections=pool_connections, pool_maxsize=pool_maxsize, max_retries=retries)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    # Some services are unstable with keep-alive; allow env var to force close (default keep-alive)
    if os.environ.get("FORCE_CONNECTION_CLOSE", "0") == "1":
        s.headers.update({"Connection": "close"})
    else:
        s.headers.update({"Connection": "keep-alive"})
    if base_headers:
        s.headers.update(base_headers)
    return s

# ---------------------------
# Extract <JSON_START> ... <JSON_END>
# ---------------------------
def extract_content(response):
    try:
        json_start = response.index("<JSON_START>")
        analysis_text = response.strip()
        json_start = json_start + len("<JSON_START>")
        json_end = response.index("<JSON_END>")
        json_str = response[json_start:json_end].strip()
        return analysis_text, json_str
    except (ValueError, json.JSONDecodeError):
        return None, None

# ---------------------------
# Reliable requests session (local VLLM)
# ---------------------------
_local_session = make_robust_session(pool_connections=300, pool_maxsize=300)

def request_local_vllm(messages, model, temperature, timeout=(10, 120), n=1):
    """
    Use local vLLM OpenAI-compatible interface, returning n choices in a single request.
    """
    url = "http://localhost:8000/v1/chat/completions"
    data = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "n": int(max(1, n)),  # ← Added: return batch choices
    }
    try:
        r = _local_session.post(url, json=data, timeout=timeout)
        r.raise_for_status()
        return r.json()
    except requests.RequestException as e:
        return {"choices": [], "error": str(e)}

# ---------------------------
# Utility: clean XLSX illegal chars & normalize cell type
# ---------------------------
_ILLEGAL_RE = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]')

def _clean_str(s: str) -> str:
    if not isinstance(s, str):
        s = str(s)
    return _ILLEGAL_RE.sub("", s)

def _coerce_cell(v):
    """Convert any value into openpyxl-acceptable content without illegal control characters."""
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, str):
        return _clean_str(v)
    # Other types → try JSON; fallback to str
    try:
        s = json.dumps(v, ensure_ascii=False)
    except TypeError:
        s = str(v)
    return _clean_str(s)

# ---------------------------
# XLSX writing (keep function name unchanged, compatible with existing calls)
# ---------------------------
def dict2csv(res, path):
    """
    Now writes to XLSX. Function name stays the same.
    - First write: create .xlsx file, write headers + row
    - Append: if new fields appear, extend header columns (old rows remain empty)
    - Rows follow header order
    - If a row contains illegal control chars causing write error → skip that row & log warning
    """
    if not res:
        return

    xlsx_filename = Path(path)
    # If user still passes .csv or other suffix, force .xlsx, to avoid confusion
    if xlsx_filename.suffix.lower() != ".xlsx":
        xlsx_filename = xlsx_filename.with_suffix(".xlsx")

    xlsx_filename.parent.mkdir(parents=True, exist_ok=True)

    try:
        if xlsx_filename.exists():
            wb = load_workbook(xlsx_filename)
            ws = wb.active

            # Read or init headers
            if ws.max_row >= 1:
                headers = [c.value for c in ws[1]]
                if headers is None:
                    headers = []
            else:
                headers = []

            # Detect new columns
            current_keys = list(res.keys())
            new_cols = [k for k in current_keys if k not in headers]
            if not headers:
                headers = current_keys
                ws.append([_coerce_cell(h) for h in headers])
            else:
                if new_cols:
                    for k in new_cols:
                        ws.cell(row=1, column=len(headers) + 1, value=_coerce_cell(k))
                        headers.append(k)

            # Write row
            row = [_coerce_cell(res.get(h, None)) for h in headers]
            try:
                ws.append(row)
                wb.save(xlsx_filename)
            except IllegalCharacterError:
                log(f"⚠️ Row skipped (contains illegal characters), userid={res.get('userid')}, sample_idx={res.get('sample_idx', 'NA')}")
                return
            except Exception as e:
                log(f"⚠️ XLSX write error (row skipped): {e}")
                return
        else:
            wb = Workbook()
            ws = wb.active
            headers = list(res.keys())
            ws.append([_coerce_cell(h) for h in headers])
            row = [_coerce_cell(res.get(h, None)) for h in headers]
            try:
                ws.append(row)
                wb.save(xlsx_filename)
            except IllegalCharacterError:
                log(f"⚠️ Illegal characters on first write, row skipped, file created. userid={res.get('userid')}")
                wb.save(xlsx_filename)
            except Exception as e:
                log(f"⚠️ First XLSX write error (row skipped): {e}")
                wb.save(xlsx_filename)
    except Exception as e:
        log(f"⚠️ Workbook processing error (row skipped): {e}")
        return

# ---------------------------
# Uplift parallel wrapper: prefer scorer.score_parallel; fallback to serial
# ---------------------------
def score_with_uplift_parallel(
    proto_uplift: ProtoKNNUpliftScorer,
    current_examples_text: str,
    response_list: list,
    current_user_input: str,
    current_userid: str,
    neighbors: int,
    coef_self: float,
    coef_neighbors: float,
    self_scores: tuple,
    parallel: bool = False,
    max_workers: int = 8,
):
    """
    Compatible return structure: same as scorer.score, containing at least:
    {ok, scores:{final/self/neighbors}, choice_idx, neighbor_count, proto_id, neighbors?}
    """
    if parallel and hasattr(proto_uplift, "score_parallel") and callable(getattr(proto_uplift, "score_parallel")):
        return proto_uplift.score_parallel(
            current_examples_text=current_examples_text,
            response_list=response_list,
            current_user_input=current_user_input,
            current_userid=current_userid,
            topk=neighbors,
            coef_self=coef_self,
            coef_neighbors=coef_neighbors,
            self_scores=self_scores,
            max_workers=max_workers,
        )

    # Fallback: serial
    return proto_uplift.score(
        current_examples_text=current_examples_text,
        response_list=response_list,
        current_user_input=current_user_input,
        current_userid=current_userid,
        topk=neighbors,
        coef_self=coef_self,
        coef_neighbors=coef_neighbors,
        self_scores=self_scores,
    )

# ---------------------------
# Call LLM (multi-sampling), average self-scores, verbose logging
# Important change: use vLLM n parameter to get multiple completions per call
# ---------------------------
def call_llm_and_prepare_data(
    temperature, userid, few_shot_examples, user_input, response_list,
    chosen_score, rejected_score, model, sample_idx: int, self_runs: int = 4, verbose: bool = False
):
    # Fix response order per sample (external ensures neighbors also use same order)
    random.shuffle(response_list)

    if verbose:
        pu = user_input.replace("\n", " ")[:120]
        r1p = response_list[0].replace("\n", " ")[:80]
        r2p = response_list[1].replace("\n", " ")[:80]
        log(f"[{userid}#{sample_idx}] ▶ Self-score start | prompt='{pu}...' | R1='{r1p}...' | R2='{r2p}...'")

    fewshots_formatted = fewshot_formatter(few_shot_examples)
    system_prompt = USER_PREFERENCE_ANALYSIS_PROMPT.format(few_shots=fewshots_formatted)
    user_prompt = USER_PROMPT_TEMPLATE.format(
        user_input=user_input,
        response_1=response_list[0],
        response_2=response_list[1],
    )
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    # Storage
    self_runs = max(0, int(self_runs))
    run_scores_r1 = [None] * self_runs
    run_scores_r2 = [None] * self_runs
    merged_examples_process = []
    merged_rationales = []
    last_examples_text = ""  # Filled by any parsed run in the batch

    def _to_float(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    # ===== Single request: n=max(1, self_runs) =====
    n_req = max(1, self_runs)
    resp = request_local_vllm(messages, model, temperature, n=n_req)
    if isinstance(resp, dict) and resp.get("error"):
        if verbose:
            log(f"[{userid}#{sample_idx}]   ✗ vLLM batch request failed: {resp.get('error')}")
    else:
        choices = resp.get("choices", [])
        analysis_taken = False

        for i, ch in enumerate(choices):
            try:
                content = ch["message"]["content"]
            except Exception:
                continue

            # Extract analysis text & JSON
            examples_analysis_process, parsed_content = extract_content(content)

            # Opportunistic: use first analysis text found for uplift
            if (examples_analysis_process is not None) and (not analysis_taken):
                merged_examples_process.append(f"[Batch] {examples_analysis_process}")
                last_examples_text = examples_analysis_process
                analysis_taken = True

            # First self_runs participate in scoring
            if i < self_runs and parsed_content:
                try:
                    parsed = json.loads(parsed_content)
                except Exception:
                    if verbose:
                        log(f"[{userid}#{sample_idx}]   ⚠ Batch {i+1} JSON parse failed")
                    continue

                llm_scores = parsed.get("better_response", {})
                llm_rationale = parsed.get("rationale", "N/A")
                if llm_rationale:
                    merged_rationales.append(f"[Run {i+1}] {llm_rationale}")

                s1 = _to_float(llm_scores.get("response_1"))
                s2 = _to_float(llm_scores.get("response_2"))
                run_scores_r1[i] = s1
                run_scores_r2[i] = s2

                if verbose:
                    log(f"[{userid}#{sample_idx}]   ✓ Batch {i+1}/{self_runs}: r1={s1}, r2={s2}")
            elif i < self_runs and not parsed_content:
                if verbose:
                    log(f"[{userid}#{sample_idx}]   ⚠ Batch {i+1}/{self_runs} no JSON extracted")

    # ===== Compute averages =====
    valid_r1 = [x for x in run_scores_r1 if x is not None]
    valid_r2 = [x for x in run_scores_r2 if x is not None]
    n1, n2 = len(valid_r1), len(valid_r2)
    self_avg_r1 = (sum(valid_r1) / n1) if n1 > 0 else None
    self_avg_r2 = (sum(valid_r2) / n2) if n2 > 0 else None

    if verbose:
        log(f"[{userid}#{sample_idx}] ▶ Self-score averages: r1={self_avg_r1}, r2={self_avg_r2} | valid counts: r1={n1}, r2={n2}")
        if not last_examples_text:
            log(f"[{userid}#{sample_idx}]   ℹ No analysis text extracted, uplift skipped (fallback to self-score)")

    # ===== Pack row =====
    res = {
        "userid": userid,
        "total_messages": json.dumps(messages, ensure_ascii=False),
        "examples_analysis_process": " | ".join(merged_examples_process) if merged_examples_process else "",
        "rationale": " | ".join(merged_rationales) if merged_rationales else "N/A",
        "chose": "",
        "response_list": json.dumps(response_list, ensure_ascii=False),
        "chosen_score": chosen_score,
        "rejected_score": rejected_score,
        "Correct/Wrong": None,
        "self_avg_response_1": self_avg_r1,
        "self_avg_response_2": self_avg_r2,
        "sample_idx": sample_idx,
    }
    for idx in range(self_runs):
        res[f"run{idx+1}_score_response_1"] = run_scores_r1[idx]
        res[f"run{idx+1}_score_response_2"] = run_scores_r2[idx]

    # Return row, average scores, and analysis text for uplift
    return res, (self_avg_r1, self_avg_r2), last_examples_text

# ---------------------------
# Concurrency control: per user
# ---------------------------
def _normalize_user_entries(user_entries):
    if isinstance(user_entries, dict) and "few_shot" in user_entries and "test_data" in user_entries:
        few = user_entries.get("few_shot", [])
        tests = user_entries.get("test_data", [])
        if isinstance(tests, list):
            return [{"few_shot": few, "test_data": t} for t in tests]
        else:
            return [{"few_shot": few, "test_data": tests}]
    if isinstance(user_entries, list):
        if all(isinstance(e, dict) and "few_shot" in e and "test_data" in e for e in user_entries):
            return user_entries
        else:
            raise ValueError("user_entries is list, but elements do not contain few_shot/test_data.")
    raise ValueError("Unknown user_entries format.")

def process_user_data_concurrently(
    temperature, userid, user_entries, csvpath, model, max_workers,
    proto_uplift: ProtoKNNUpliftScorer, neighbors: int, coef_self: float, coef_neighbors: float,
    self_runs: int = 4, verbose: bool = False,
    uplift_parallel: bool = False, uplift_workers: int = 8,
):
    user_entries = _normalize_user_entries(user_entries)
    written = 0
    correct = 0

    def submit_one(sample_idx, entry):
        # Slight jitter to avoid sudden spikes
        time.sleep(random.uniform(0, 0.2))

        few_shot_examples = entry.get("few_shot", [])
        test_item = entry.get("test_data", None)
        if not test_item:
            return None
        user_input = test_item["context"][0]["content"]
        chosen_text = test_item["chosen"]["content"]       # Gold answer (text)
        rejected_text = test_item["rejected"]["content"]
        responses_list = [chosen_text, rejected_text]      # This gets shuffled; gold remains chosen_text
        raw_chosen = test_item.get("chosen_score", None)
        raw_rejected = test_item.get("rejected_score", None)
        chosen_score = (raw_chosen / 10) if isinstance(raw_chosen, (int, float)) else None
        rejected_score = (raw_rejected / 10) if isinstance(raw_rejected, (int, float)) else None

        res_row, self_scores, last_examples_text = call_llm_and_prepare_data(
            temperature, userid, few_shot_examples, user_input, responses_list,
            chosen_score, rejected_score, model, sample_idx=sample_idx, self_runs=self_runs, verbose=verbose
        )
        # Store gold answer
        res_row["golden_chosen"] = chosen_text

        uplift_ok = False
        final_choice_idx = None
        final_score_r1 = None
        final_score_r2 = None
        decision_source = None  # "uplift" or "self_avg" / "self_fallback"

        if last_examples_text:
            # In uplift-only mode: self-score weight = 0, neighbors weight = 1, and pass unpackable 2-tuple
            if self_runs == 0:
                eff_coef_self = 0.0
                eff_coef_neighbors = 1.0
                eff_self_scores = (0.0, 0.0)  # ★ Avoid None unpacking error
            else:
                eff_coef_self = coef_self
                eff_coef_neighbors = coef_neighbors
                eff_self_scores = self_scores

            # Uplift call with safety; avoid crashing thread due to network noise
            try:
                uplift = score_with_uplift_parallel(
                    proto_uplift=proto_uplift,
                    current_examples_text=last_examples_text,     # Internally extracts history block
                    response_list=responses_list,                # Order preserved
                    current_user_input=user_input,
                    current_userid=userid,
                    neighbors=neighbors,
                    coef_self=eff_coef_self,
                    coef_neighbors=eff_coef_neighbors,
                    self_scores=eff_self_scores,
                    parallel=uplift_parallel,
                    max_workers=uplift_workers,
                )
            except Exception as e:
                uplift = {"ok": False}
                if verbose:
                    log(f"[{userid}#{sample_idx}] ▶ Uplift call exception: {e} (fallback to self-score or skip)")

            if uplift.get("ok"):
                uplift_ok = True
                res_row["uplift_proto_id"] = uplift.get("proto_id")
                res_row["uplift_neighbor_count"] = uplift.get("neighbor_count")
                res_row["uplift_final_score_r1"] = uplift["scores"]["final"]["r1"]
                res_row["uplift_final_score_r2"] = uplift["scores"]["final"]["r2"]
                res_row["uplift_self_r1"] = uplift["scores"]["self"].get("r1")
                res_row["uplift_self_r2"] = uplift["scores"]["self"].get("r2")
                res_row["uplift_nb_r1_avg"] = uplift["scores"]["neighbors"].get("r1")
                res_row["uplift_nb_r2_avg"] = uplift["scores"]["neighbors"].get("r2")
                final_score_r1 = uplift["scores"]["final"]["r1"]
                final_score_r2 = uplift["scores"]["final"]["r2"]
                final_choice_idx = uplift["choice_idx"]
                decision_source = "uplift"
                res_row["uplift_choice"] = responses_list[final_choice_idx]

                if verbose:
                    nlist = uplift.get("neighbors", [])
                    log(f"[{userid}#{sample_idx}] ▶ Uplift success | proto={uplift.get('proto_id')} | topk={uplift.get('neighbor_count')} | coef_self={eff_coef_self}, coef_neighbors={eff_coef_neighbors}")
                    log(f"[{userid}#{sample_idx}]   Self-score averages: r1={res_row['uplift_self_r1']}, r2={res_row['uplift_self_r2']}")
                    log(f"[{userid}#{sample_idx}]   Neighbor averages:   r1={res_row['uplift_nb_r1_avg']},  r2={res_row['uplift_nb_r2_avg']}")
                    log(f"[{userid}#{sample_idx}]   Fusion result:       r1={final_score_r1}, r2={final_score_r2} → choose: {'R1' if final_choice_idx==0 else 'R2'}")
                    if nlist:
                        brief = ", ".join([str(n.get("row_index","?")) for n in nlist[:5]])
                        log(f"[{userid}#{sample_idx}]   Neighbor examples (row_index): {brief}{' ...' if len(nlist)>5 else ''}")
            else:
                res_row["uplift_choice"] = ""
                if verbose:
                    log(f"[{userid}#{sample_idx}] ▶ Uplift unavailable (no history / empty prototype / neighbor scoring failure etc.)")

        # If no uplift, fall back to self-score; if uplift used, use its choice_idx.
        if uplift_ok:
            res_row["chose"] = res_row["uplift_choice"]
        else:
            # Uplift-only mode: perform 1 lightweight fallback self-score; skip if fails
            if self_runs == 0:
                if verbose:
                    log(f"[{userid}#{sample_idx}] ▶ Uplift-only fallback: performing 1 lightweight self-score")
                try:
                    fb_row, fb_scores, _ = call_llm_and_prepare_data(
                        temperature, userid, few_shot_examples, user_input, responses_list,
                        chosen_score, rejected_score, model, sample_idx=sample_idx, self_runs=1, verbose=verbose
                    )
                    s1, s2 = fb_scores
                except Exception as e:
                    if verbose:
                        log(f"[{userid}#{sample_idx}] ✗ Fallback self-score error: {e}")
                    return None

                if s1 is not None and s2 is not None:
                    final_choice_idx = 0 if s1 >= s2 else 1
                    final_score_r1, final_score_r2 = s1, s2
                    decision_source = "self_fallback"
                    res_row["chose"] = responses_list[final_choice_idx]
                    if verbose:
                        who = "R1" if final_choice_idx==0 else "R2"
                        log(f"[{userid}#{sample_idx}] ▶ Fallback self-score decision → choose: {who} (r1={s1}, r2={s2})")
                else:
                    if verbose:
                        log(f"[{userid}#{sample_idx}] ✗ Fallback self-score invalid, skipping write")
                    return None
            else:
                s1, s2 = self_scores
                if s1 is not None and s2 is not None:
                    final_choice_idx = 0 if s1 > s2 else (1 if s2 > s1 else 0)  # tie → default 0
                    final_score_r1, final_score_r2 = s1, s2
                    decision_source = "self_avg"
                    res_row["chose"] = responses_list[final_choice_idx]
                    if verbose:
                        who = "R1" if final_choice_idx==0 else "R2"
                        log(f"[{userid}#{sample_idx}] ▶ Self-score decision → choose: {who} (r1={s1}, r2={s2})")
                else:
                    if verbose:
                        log(f"[{userid}#{sample_idx}] ✗ Neither self-score nor uplift available, skipping write")
                    return None

        # Fill final scores & decision source
        res_row["final_score_r1"] = final_score_r1
        res_row["final_score_r2"] = final_score_r2
        res_row["decision_source"] = decision_source

        # Compute correctness: compare final choice text vs gold
        is_correct = (res_row["chose"] == chosen_text)
        res_row["Correct/Wrong"] = bool(is_correct)

        if verbose:
            gold_idx = 0 if chosen_text == responses_list[0] else 1
            gold_tag = "R1" if gold_idx==0 else "R2"
            final_tag = "R1" if final_choice_idx==0 else "R2"
            log(f"[{userid}#{sample_idx}] ✓ Final choice: {final_tag} | GOLD={gold_tag} → {'✔ correct' if is_correct else '✘ wrong'}")

        return res_row
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = [ex.submit(submit_one, idx, entry) for idx, entry in enumerate(user_entries)]
        for fut in tqdm(concurrent.futures.as_completed(futs), total=len(futs), desc=f"Processing User {userid}"):
            try:
                out = fut.result()
            except Exception as e:
                log(f"⚠️ Thread execution error(userid={userid}): {e}")
                continue
            if out:
                dict2csv(out, csvpath)
                written += 1
                if out.get("Correct/Wrong") is True:
                    correct += 1

    # Return written count and correct count for this user
    return written, correct

# ---------------------------
# Main
# ---------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Score after GRPO with prototype KNN uplift (averages).")
    parser.add_argument("--model", type=str, required=True, help="Model name for local vLLM service")
    parser.add_argument("--fs", type=int, required=True, help="few-shot count (for data split)")
    parser.add_argument("--max-workers", type=int, default=36)
    parser.add_argument("--dataset", type=str, required=True, choices=["prism", "chatbot"])
    parser.add_argument("--space", type=str, required=True)
    parser.add_argument("--verbose", action="store_true", help="Print intermediate process for each sample")

    # Uplift / prototype related
    parser.add_argument("--proto-save-dir", type=str, required=True, help="Training output directory (containing E.pt/A*.pt/assign*.pt/row_index.pt)")
    parser.add_argument("--st-embed", type=str, required=True, help="Local sentence-transformers model path used during training")
    parser.add_argument("--dataset-csv", type=str, required=True, help="Original CSV path used for training/clustering")
    parser.add_argument("--examples-col", type=str, default="examples_analysis_process")
    parser.add_argument("--neighbors", type=int, default=4)
    parser.add_argument("--coef-self", type=float, default=0.5)
    parser.add_argument("--coef-neighbors", type=float, default=0.5)
    parser.add_argument("--use-trained", action="store_true")
    parser.add_argument("--temperature", type=float, required=True)

    # Added: self-score repeat count (previously fixed at 4)
    parser.add_argument("--self-runs", type=int, default=4, help="Number of repeated self-score runs for each sample")

    # Added: uplift parallel control
    parser.add_argument("--uplift-parallel", action="store_true", help="Parallel neighbor preference fusion (if scorer supports; otherwise fallback to serial)")
    parser.add_argument("--uplift-workers", type=int, default=8, help="Neighbor parallel thread count (only effective when --uplift-parallel is used)")

    args = parser.parse_args()

    MODEL_NAME = args.model
    MAX_WORKERS = args.max_workers
    NUM_FEWSHOTS = args.fs
    NUM_RUNS = 1
    space = args.space
    temperature = args.temperature
    self_runs = args.self_runs

    # Dataset
    print("Loading dataset...")
    if args.dataset == "prism":
        dataset = Prism_personal_align_Dataset()
    elif args.dataset == "chatbot":
        dataset = ChatbotArenaDataset()
    else:
        raise ValueError("Unknown dataset")
    all_data = dataset.get_all_user_data()
    print("Dataset loaded.")

    # Uplift scorer
    proto_uplift = ProtoKNNUpliftScorer(
        save_dir=args.proto_save_dir,
        csv_path=args.dataset_csv,
        examples_col=args.examples_col,
        embed_model_path=args.st_embed,
        device="cuda:0",
        use_trained=args.use_trained,
        llm_model_name="gpt-4o-2024-11-20"
    )

    # Use xlsx output
    BASE_OUTPUT_PATH = f"/mnt/{space}/zpy/Pers_RM/vllm_pre_experiments/with_proto/{MODEL_NAME}.xlsx"

    # Global stats
    global_total = 0
    global_correct = 0

    for i in range(NUM_RUNS):
        run_num = i + 1
        output_csv = BASE_OUTPUT_PATH.replace(".xlsx", f"_run_{run_num}_with_uplift.xlsx")
        print(f"\n--- Starting run {run_num}/{NUM_RUNS} ---")
        random.seed(42)
        gft = get_full_train_val_rotate(all_data, n_examples=NUM_FEWSHOTS, test_split_name="test")

        # Stats for this run
        run_total = 0
        run_correct = 0

        for userid, data in gft.items():
            written, correct = process_user_data_concurrently(
                temperature, userid, data, csvpath=output_csv, model=MODEL_NAME, max_workers=MAX_WORKERS,
                proto_uplift=proto_uplift, neighbors=args.neighbors,
                coef_self=args.coef_self, coef_neighbors=args.coef_neighbors,
                self_runs=self_runs, verbose=args.verbose,
                uplift_parallel=args.uplift_parallel, uplift_workers=args.uplift_workers,
            )
            run_total += written
            run_correct += correct

            # Accuracy per user
            user_acc = (correct / written * 100.0) if written > 0 else 0.0
            print(f"User {userid} stats: written {written}, correct {correct}, accuracy {user_acc:.2f}%")

        # Accuracy for this run
        run_acc = (run_correct / run_total * 100.0) if run_total > 0 else 0.0
        print(f"\n== Run {run_num} summary ==")
        print(f"Samples written: {run_total}")
        print(f"Correct samples: {run_correct}")
        print(f"Accuracy: {run_acc:.2f}%")
        print(f"Results saved to: {output_csv}")

        global_total += run_total
        global_correct += run_correct

    # Global accuracy
    global_acc = (global_correct / global_total * 100.0) if global_total > 0 else 0.0
    print("\n================ Global Summary ================")
    print(f"Total samples: {global_total}")
    print(f"Total correct: {global_correct}")
    print(f"Global accuracy: {global_acc:.2f}%")
    print("===============================================")
