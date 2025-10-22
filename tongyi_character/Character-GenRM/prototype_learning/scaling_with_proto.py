# -*- coding: utf-8 -*-
import json
import random
import csv
from pathlib import Path
import argparse
import concurrent.futures
import threading
import re  
import time  

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from tqdm import tqdm
import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from pre_process.dataset import PrismDataset, ChatbotArenaDataset,Prism_personal_align_Dataset
from proto_knn_scaler import ProtoKNNUpliftScorer
from inference.PROMPTS import USER_PROMPT_TEMPLATE,USER_PREFERENCE_ANALYSIS_PROMPT
from utils import fewshot_formatter, get_full_train_val_rotate

# Added: xlsx writing dependency
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError  # ← Added: exception type

# ---------------------------
# Thread-safe print
# ---------------------------
_print_lock = threading.Lock()
def log(msg: str):
    with _print_lock:
        print(msg, flush=True)

# ---------------------------
# Robust HTTP Session (POST also retries)
# ---------------------------
def make_robust_session(
    total=5,
    connect=5,
    read=5,
    status=5,
    backoff_factor=0.5,
    status_forcelist=(429, 500, 502, 503, 504),
    allowed_methods=frozenset(['GET', 'POST', 'PUT', 'DELETE', 'HEAD', 'OPTIONS', 'TRACE', 'PATCH']),
    pool_connections=200,
    pool_maxsize=200,
    base_headers=None,
):
    s = requests.Session()
    retries = Retry(
        total=total,
        connect=connect,
        read=read,
        status=status,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
        allowed_methods=allowed_methods,
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(pool_connections=pool_connections, pool_maxsize=pool_maxsize, max_retries=retries)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    # Some services are unstable with keep-alive; you can force short connections via env var (default keep-alive)
    if os.environ.get("FORCE_CONNECTION_CLOSE", "0") == "1":
        s.headers.update({"Connection": "close"})
    else:
        s.headers.update({"Connection": "keep-alive"})
    if base_headers:
        s.headers.update(base_headers)
    return s

# ---------------------------
# Extract <JSON_START> ... <JSON_END>
# ---------------------------
def extract_content(response):
    try:
        json_start = response.index("<JSON_START>")
        analysis_text = response.strip()
        json_start = json_start + len("<JSON_START>")
        json_end = response.index("<JSON_END>")
        json_str = response[json_start:json_end].strip()
        return analysis_text, json_str
    except (ValueError, json.JSONDecodeError):
        return None, None

# ---------------------------
# Reliable requests session (local VLLM)
# ---------------------------
_local_session = make_robust_session(pool_connections=300, pool_maxsize=300)

def request_local_vllm(messages, model, temperature, timeout=(10, 120)):
    url = "http://localhost:8000/v1/chat/completions"
    data = {"model": model, "messages": messages, "temperature": temperature}
    try:
        r = _local_session.post(url, json=data, timeout=timeout)
        r.raise_for_status()
        return r.json()
    except requests.RequestException as e:
        return {"choices": [{"message": {"content": ""}}], "error": str(e)}

# ---------------------------
# Utility: clean XLSX illegal characters & normalize cell types
# ---------------------------
_ILLEGAL_RE = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]')

def _clean_str(s: str) -> str:
    if not isinstance(s, str):
        s = str(s)
    return _ILLEGAL_RE.sub("", s)

def _coerce_cell(v):
    """Convert any value to an openpyxl-acceptable form without illegal control characters."""
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, str):
        return _clean_str(v)
    # Other types: try to convert to JSON; if fails, convert to str
    try:
        s = json.dumps(v, ensure_ascii=False)
    except TypeError:
        s = str(v)
    return _clean_str(s)

# ---------------------------
# XLSX writing (keep function name unchanged for compatibility)
# ---------------------------
def dict2csv(res, path):
    """
    Now writes to XLSX. Keeps the same function name and call style.
    - First write: create .xlsx file, write header and data
    - Append: if new fields appear, automatically extend header columns (old rows remain blank)
    - Row data written in header order
    - If a row contains illegal control characters causing write errors, skip that row and log
    """
    if not res:
        return

    xlsx_filename = Path(path)
    # If suffix is still .csv etc., force change to .xlsx to avoid confusion
    if xlsx_filename.suffix.lower() != ".xlsx":
        xlsx_filename = xlsx_filename.with_suffix(".xlsx")

    xlsx_filename.parent.mkdir(parents=True, exist_ok=True)

    try:
        if xlsx_filename.exists():
            wb = load_workbook(xlsx_filename)
            ws = wb.active

            # Read or initialize header
            if ws.max_row >= 1:
                headers = [c.value for c in ws[1]]
                if headers is None:
                    headers = []
            else:
                headers = []

            # Extend header if new fields found
            current_keys = list(res.keys())
            new_cols = [k for k in current_keys if k not in headers]
            if not headers:
                headers = current_keys
                # Clean header names to prevent unexpected keys
                ws.append([_coerce_cell(h) for h in headers])
            else:
                if new_cols:
                    for k in new_cols:
                        ws.cell(row=1, column=len(headers) + 1, value=_coerce_cell(k))
                        headers.append(k)

            # Write one row (clean each cell)
            row = [_coerce_cell(res.get(h, None)) for h in headers]
            try:
                ws.append(row)
                wb.save(xlsx_filename)
            except IllegalCharacterError:
                log(f"⚠️ Skipped one row (contains illegal character), userid={res.get('userid')}, sample_idx={res.get('sample_idx', 'NA')}")
                return
            except Exception as e:
                log(f"⚠️ XLSX write error (row skipped): {e}")
                return
        else:
            wb = Workbook()
            ws = wb.active
            headers = list(res.keys())
            ws.append([_coerce_cell(h) for h in headers])
            row = [_coerce_cell(res.get(h, None)) for h in headers]
            try:
                ws.append(row)
                wb.save(xlsx_filename)
            except IllegalCharacterError:
                log(f"⚠️ Illegal character in first write, row skipped, file created. userid={res.get('userid')}")
                wb.save(xlsx_filename)
            except Exception as e:
                log(f"⚠️ XLSX first write error (row skipped): {e}")
                wb.save(xlsx_filename)
    except Exception as e:
        # Other exceptions (e.g., concurrent access) should not block the process
        log(f"⚠️ Workbook handling error (row skipped): {e}")
        return

def call_llm_and_prepare_data(
    temperature, userid, few_shot_examples, user_input, response_list,
    chosen_score, rejected_score, model, sample_idx: int, self_runs: int = 4, verbose: bool = False
):
    # Fix the response order for each sample (make sure the neighbors side uses the same order)
    random.shuffle(response_list)

    if verbose:
        pu = user_input.replace("\n", " ")[:120]
        r1p = response_list[0].replace("\n", " ")[:80]
        r2p = response_list[1].replace("\n", " ")[:80]
        log(f"[{userid}#{sample_idx}] ▶ Self-scoring starts | prompt='{pu}...' | R1='{r1p}...' | R2='{r2p}...'")

    fewshots_formatted = fewshot_formatter(few_shot_examples)
    system_prompt = USER_PREFERENCE_ANALYSIS_PROMPT.format(few_shots=fewshots_formatted)
    user_prompt = USER_PROMPT_TEMPLATE.format(
        user_input=user_input,
        response_1=response_list[0],
        response_2=response_list[1],
    )
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    # Configurable number of scoring runs
    self_runs = max(0, int(self_runs))
    run_scores_r1 = [None] * self_runs
    run_scores_r2 = [None] * self_runs
    merged_examples_process = []
    merged_rationales = []
    last_examples_text = ""  # for uplift history parsing
    if self_runs == 0:
        try:
            response0 = request_local_vllm(messages, model, temperature)
            if isinstance(response0, dict) and "error" in response0:
                if verbose:
                    log(f"[{userid}#{sample_idx}]   ✗ Run 0 LLM request failed: {response0.get('error')}")
            else:
                try:
                    content0 = response0["choices"][0]["message"]["content"]
                    examples_analysis_process0, parsed_content0 = extract_content(content0)
                    if examples_analysis_process0:
                        merged_examples_process.append(f"[Run 0] {examples_analysis_process0}")
                        last_examples_text = examples_analysis_process0  # for uplift use
                    if parsed_content0:
                        parsed0 = json.loads(parsed_content0)
                        llm_rationale0 = parsed0.get("rationale", "N/A")
                        if llm_rationale0:
                            merged_rationales.append(f"[Run 0] {llm_rationale0}")
                    if verbose:
                        log(f"[{userid}#{sample_idx}]   ✓ Run 0 extraction succeeded (uplift only, not scored)")
                except (KeyError, IndexError, TypeError, json.JSONDecodeError) as e:
                    if verbose:
                        log(f"[{userid}#{sample_idx}]   ⚠ Run 0 parsing failed: {e}")
        except Exception as e:
            if verbose:
                log(f"[{userid}#{sample_idx}]   ⚠ Run 0 exception: {e}")

    def _to_float(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    for i in range(self_runs):
        response = request_local_vllm(messages, model, temperature)
        if isinstance(response, dict) and "error" in response:
            if verbose:
                log(f"[{userid}#{sample_idx}]   ✗ {i+1}/{self_runs}th LLM request failed: {response.get('error')}")
            continue
        try:
            content = response["choices"][0]["message"]["content"]
        except (KeyError, IndexError, TypeError):
            if verbose:
                log(f"[{userid}#{sample_idx}]   ✗ {i+1}/{self_runs}th response parsing failed (no content)")
            continue

        try:
            examples_analysis_process, parsed_content = extract_content(content)
            if examples_analysis_process:
                merged_examples_process.append(f"[Run {i+1}] {examples_analysis_process}")
                last_examples_text = examples_analysis_process  # for uplift (take from current LLM output)

            if not parsed_content:
                if verbose:
                    log(f"[{userid}#{sample_idx}]   ⚠ {i+1}/{self_runs}th no JSON parsed")
                continue

            parsed = json.loads(parsed_content)
            llm_scores = parsed.get("better_response", {})
            llm_rationale = parsed.get("rationale", "N/A")
            if llm_rationale:
                merged_rationales.append(f"[Run {i+1}] {llm_rationale}")

            s1 = _to_float(llm_scores.get("response_1"))
            s2 = _to_float(llm_scores.get("response_2"))
            run_scores_r1[i] = s1 if s1 is not None else None
            run_scores_r2[i] = s2 if s2 is not None else None

            if verbose:
                log(f"[{userid}#{sample_idx}]   ✓ {i+1}/{self_runs}th parsed successfully: r1={s1}, r2={s2}")

        except (json.JSONDecodeError, TypeError) as e:
            if verbose:
                log(f"[{userid}#{sample_idx}]   ✗ {i+1}/{self_runs}th JSON parsing exception: {e}")
            continue

    # Compute the 'average score' (count successful runs only)
    valid_r1 = [x for x in run_scores_r1 if x is not None]
    valid_r2 = [x for x in run_scores_r2 if x is not None]
    n1, n2 = len(valid_r1), len(valid_r2)

    self_avg_r1 = (sum(valid_r1) / n1) if n1 > 0 else None
    self_avg_r2 = (sum(valid_r2) / n2) if n2 > 0 else None

    if verbose:
        log(f"[{userid}#{sample_idx}] ▶ Self-score means: r1={self_avg_r1}, r2={self_avg_r2} | valid runs: r1={n1}, r2={n2}")

    # Assemble base XLSX row (dynamic run fields)
    res = {
        "userid": userid,
        "total_messages": json.dumps(messages, ensure_ascii=False),
        "examples_analysis_process": " | ".join(merged_examples_process) if merged_examples_process else "",
        "rationale": " | ".join(merged_rationales) if merged_rationales else "N/A",
        "chose": "",  # leave empty; filled later by uplift or self-score decision
        "response_list": json.dumps(response_list, ensure_ascii=False),
        "chosen_score": chosen_score,
        "rejected_score": rejected_score,
        "Correct/Wrong": None,  # fill when a golden answer is available
        # Mean (self-scoring)
        "self_avg_response_1": self_avg_r1,
        "self_avg_response_2": self_avg_r2,
    }
    # Dynamically add the score of each run
    for idx in range(self_runs):
        res[f"run{idx+1}_score_response_1"] = run_scores_r1[idx]
        res[f"run{idx+1}_score_response_2"] = run_scores_r2[idx]

    # For logging: also include sample index (may be used when writing files)
    res["sample_idx"] = sample_idx

    return res, (self_avg_r1, self_avg_r2), last_examples_text

# ---------------------------
# Concurrency control: per user
# ---------------------------
def _normalize_user_entries(user_entries):
    if isinstance(user_entries, dict) and "few_shot" in user_entries and "test_data" in user_entries:
        few = user_entries.get("few_shot", [])
        tests = user_entries.get("test_data", [])
        if isinstance(tests, list):
            return [{"few_shot": few, "test_data": t} for t in tests]
        else:
            return [{"few_shot": few, "test_data": tests}]
    if isinstance(user_entries, list):
        if all(isinstance(e, dict) and "few_shot" in e and "test_data" in e for e in user_entries):
            return user_entries
        else:
            raise ValueError("user_entries is a list, but elements lack few_shot/test_data.")
    raise ValueError("Unknown user_entries format.")

def process_user_data_concurrently(
    temperature, userid, user_entries, csvpath, model, max_workers,
    proto_uplift: ProtoKNNUpliftScorer, neighbors: int, coef_self: float, coef_neighbors: float,
    self_runs: int = 4, verbose: bool = False
):
    user_entries = _normalize_user_entries(user_entries)
    written = 0
    correct = 0

    def submit_one(sample_idx, entry):
        # Small jitter to reduce burst
        time.sleep(random.uniform(0, 0.2))

        few_shot_examples = entry.get("few_shot", [])
        test_item = entry.get("test_data", None)
        if not test_item:
            return None
        user_input = test_item["context"][0]["content"]
        chosen_text = test_item["chosen"]["content"]       # golden answer (text)
        rejected_text = test_item["rejected"]["content"]
        responses_list = [chosen_text, rejected_text]      # will be shuffled again; gold is still chosen_text
        raw_chosen = test_item.get("chosen_score", None)
        raw_rejected = test_item.get("rejected_score", None)
        chosen_score = (raw_chosen / 10) if isinstance(raw_chosen, (int, float)) else None
        rejected_score = (raw_rejected / 10) if isinstance(raw_rejected, (int, float)) else None

        res_row, self_scores, last_examples_text = call_llm_and_prepare_data(
            temperature, userid, few_shot_examples, user_input, responses_list,
            chosen_score, rejected_score, model, sample_idx=sample_idx, self_runs=self_runs, verbose=verbose
        )
        # Record golden answer correctly
        res_row["golden_chosen"] = chosen_text

        uplift_ok = False
        final_choice_idx = None
        final_score_r1 = None
        final_score_r2 = None
        decision_source = None  # "uplift" or "self_avg" / "self_fallback"

        if last_examples_text:
            # Uplift-only mode: set self-score weight to 0, neighbor weight to 1, and pass an unpackable tuple
            if self_runs == 0:
                eff_coef_self = 0.0
                eff_coef_neighbors = 1.0
                eff_self_scores = (0.0, 0.0)  # ★ avoid unpacking error from None
            else:
                eff_coef_self = coef_self
                eff_coef_neighbors = coef_neighbors
                eff_self_scores = self_scores

            # Uplift call with guard, avoid thread crash due to network jitters
            try:
                uplift = proto_uplift.score(
                    current_examples_text=last_examples_text,     # internally extracts history block
                    response_list=responses_list,                # keep the same order
                    current_user_input=user_input,
                    current_userid=userid,
                    topk=neighbors,
                    coef_self=eff_coef_self,
                    coef_neighbors=eff_coef_neighbors,
                    self_scores=eff_self_scores,
                )
            except Exception as e:
                uplift = {"ok": False}
                if verbose:
                    log(f"[{userid}#{sample_idx}] ▶ Uplift call exception: {e} (will fall back to self-scoring or skip)")

            if uplift.get("ok"):
                uplift_ok = True
                res_row["uplift_proto_id"] = uplift.get("proto_id")
                res_row["uplift_neighbor_count"] = uplift.get("neighbor_count")
                res_row["uplift_final_score_r1"] = uplift["scores"]["final"]["r1"]
                res_row["uplift_final_score_r2"] = uplift["scores"]["final"]["r2"]
                res_row["uplift_self_r1"] = uplift["scores"]["self"].get("r1")
                res_row["uplift_self_r2"] = uplift["scores"]["self"].get("r2")
                res_row["uplift_nb_r1_avg"] = uplift["scores"]["neighbors"].get("r1")
                res_row["uplift_nb_r2_avg"] = uplift["scores"]["neighbors"].get("r2")
                final_score_r1 = uplift["scores"]["final"]["r1"]
                final_score_r2 = uplift["scores"]["final"]["r2"]
                final_choice_idx = uplift["choice_idx"]
                decision_source = "uplift"
                res_row["uplift_choice"] = responses_list[final_choice_idx]

                if verbose:
                    nlist = uplift.get("neighbors", [])
                    log(f"[{userid}#{sample_idx}] ▶ Uplift succeeded | proto={uplift.get('proto_id')} | topk={uplift.get('neighbor_count')} | coef_self={eff_coef_self}, coef_neighbors={eff_coef_neighbors}")
                    log(f"[{userid}#{sample_idx}]   Self-score means: r1={res_row['uplift_self_r1']}, r2={res_row['uplift_self_r2']}")
                    log(f"[{userid}#{sample_idx}]   Neighbor means: r1={res_row['uplift_nb_r1_avg']},  r2={res_row['uplift_nb_r2_avg']}")
                    log(f"[{userid}#{sample_idx}]   Fused result: r1={final_score_r1}, r2={final_score_r2} → choose: {'R1' if final_choice_idx==0 else 'R2'}")
                    if nlist:
                        brief = ", ".join([str(n.get("row_index","?")) for n in nlist[:5]])
                        log(f"[{userid}#{sample_idx}]   Neighbor examples (row_index): {brief}{' ...' if len(nlist)>5 else ''}")
            else:
                res_row["uplift_choice"] = ""
                if verbose:
                    log(f"[{userid}#{sample_idx}] ▶ Uplift unavailable (no history / prototype has no members / neighbor scoring failed, etc.)")

        # If no uplift, compare self-scores; if uplift present, use its choice_idx.
        if uplift_ok:
            res_row["chose"] = res_row["uplift_choice"]
        else:
            # In uplift-only mode, try one lightweight self-scoring fallback; skip if it fails
            if self_runs == 0:
                if verbose:
                    log(f"[{userid}#{sample_idx}] ▶ uplift-only mode fallback: perform 1 lightweight self-scoring")
                try:
                    fb_row, fb_scores, _ = call_llm_and_prepare_data(
                        temperature, userid, few_shot_examples, user_input, responses_list,
                        chosen_score, rejected_score, model, sample_idx=sample_idx, self_runs=1, verbose=verbose
                    )
                    s1, s2 = fb_scores
                except Exception as e:
                    if verbose:
                        log(f"[{userid}#{sample_idx}] ✗ Fallback self-scoring exception: {e}")
                    return None

                if s1 is not None and s2 is not None:
                    final_choice_idx = 0 if s1 >= s2 else 1
                    final_score_r1, final_score_r2 = s1, s2
                    decision_source = "self_fallback"
                    res_row["chose"] = responses_list[final_choice_idx]
                    if verbose:
                        who = "R1" if final_choice_idx==0 else "R2"
                        log(f"[{userid}#{sample_idx}] ▶ Fallback self-scoring decision → choose: {who} (r1={s1}, r2={s2})")
                else:
                    if verbose:
                        log(f"[{userid}#{sample_idx}] ✗ Fallback self-scoring invalid, skip writing")
                    return None
            else:
                s1, s2 = self_scores
                if s1 is not None and s2 is not None:
                    final_choice_idx = 0 if s1 > s2 else (1 if s2 > s1 else 0)  # default 0 when tie
                    final_score_r1, final_score_r2 = s1, s2
                    decision_source = "self_avg"
                    res_row["chose"] = responses_list[final_choice_idx]
                    if verbose:
                        who = "R1" if final_choice_idx==0 else "R2"
                        log(f"[{userid}#{sample_idx}] ▶ Self-scoring only decision → choose: {who} (r1={s1}, r2={s2})")
                else:
                    if verbose:
                        log(f"[{userid}#{sample_idx}] ✗ Both self-scoring and uplift unavailable, skip writing")
                    return None

        # Fill back final scores & decision source
        res_row["final_score_r1"] = final_score_r1
        res_row["final_score_r2"] = final_score_r2
        res_row["decision_source"] = decision_source

        # Compute correctness: compare final choice text vs golden text
        is_correct = (res_row["chose"] == chosen_text)
        res_row["Correct/Wrong"] = bool(is_correct)

        if verbose:
            gold_idx = 0 if chosen_text == responses_list[0] else 1
            gold_tag = "R1" if gold_idx==0 else "R2"
            final_tag = "R1" if final_choice_idx==0 else "R2"
            log(f"[{userid}#{sample_idx}] ✓ Final choice: {final_tag} | GOLD={gold_tag} → {'✔ Correct' if is_correct else '✘ Wrong'}")

        return res_row

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = [ex.submit(submit_one, idx, entry) for idx, entry in enumerate(user_entries)]
        for fut in tqdm(concurrent.futures.as_completed(futs), total=len(futs), desc=f"Processing User {userid}"):
            try:
                out = fut.result()
            except Exception as e:
                log(f"⚠️ Thread execution exception (userid={userid}): {e}")
                continue
            if out:
                dict2csv(out, csvpath)
                written += 1
                if out.get("Correct/Wrong") is True:
                    correct += 1

    # Return number written and number correct for this user
    return written, correct

# ---------------------------
# Main
# ---------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Score after GRPO with prototype KNN uplift (averages).")
    parser.add_argument("--model", type=str, required=True, help="Model name for local vLLM service")
    parser.add_argument("--fs", type=int, required=True, help="Number of few-shots (for data split)")
    parser.add_argument("--max-workers", type=int, default=36)
    parser.add_argument("--dataset", type=str, required=True, choices=["prism", "chatbot"])   
    parser.add_argument("--verbose", action="store_true", help="Print intermediate process for each sample")

    # Uplift / prototype related
    parser.add_argument("--proto-save-dir", type=str, required=True, help="Training output directory (contains E.pt/A*.pt/assign*.pt/row_index.pt)")
    parser.add_argument("--st-embed", type=str, required=True, help="Path to local sentence-transformers model used during training")
    parser.add_argument("--dataset-csv", type=str, required=True, help="Path to original CSV for training/clustering")
    parser.add_argument("--examples-col", type=str, default="examples_analysis_process")
    parser.add_argument("--neighbors", type=int, default=4)
    parser.add_argument("--coef-self", type=float, default=0.5)
    parser.add_argument("--coef-neighbors", type=float, default=0.5)
    parser.add_argument("--use-trained", action="store_true")
    parser.add_argument("--temperature", type=float, required=True)

    # New: self-scoring repeat count (was hardcoded to 4)
    parser.add_argument("--self-runs", type=int, default=4, help="Number of repeated self-scoring runs per sample")

    args = parser.parse_args()

    MODEL_NAME = args.model
    MAX_WORKERS = args.max_workers
    NUM_FEWSHOTS = args.fs
    NUM_RUNS = 1
    temperature = args.temperature
    self_runs = args.self_runs

    # Dataset
    print("Loading dataset...")
    if args.dataset == "prism":
        dataset = Prism_personal_align_Dataset()
    elif args.dataset == "chatbot":
        dataset = ChatbotArenaDataset()
    else:
        raise ValueError("Unknown dataset")
    all_data = dataset.get_all_user_data()
    print("Dataset loaded.")

    # Uplift scorer
    proto_uplift = ProtoKNNUpliftScorer(
        save_dir=args.proto_save_dir,
        csv_path=args.dataset_csv,
        examples_col=args.examples_col,
        embed_model_path=args.st_embed,
        device="cuda:0",
        use_trained=args.use_trained,
    )

    # Use xlsx output
    BASE_OUTPUT_PATH = f"../vllm_experiments/with_proto/{MODEL_NAME}.xlsx"

    # Global stats
    global_total = 0
    global_correct = 0

    for i in range(NUM_RUNS):
        run_num = i + 1
        output_csv = BASE_OUTPUT_PATH.replace(".xlsx", f"_run_{run_num}_with_uplift.xlsx")
        print(f"\n--- Starting run {run_num}/{NUM_RUNS} ---")
        random.seed(42)
        gft = get_full_train_val_rotate(all_data, n_examples=NUM_FEWSHOTS, test_split_name="test")

        # This run stats
        run_total = 0
        run_correct = 0

        for userid, data in gft.items():
            written, correct = process_user_data_concurrently(
                temperature, userid, data, csvpath=output_csv, model=MODEL_NAME, max_workers=MAX_WORKERS,
                proto_uplift=proto_uplift, neighbors=args.neighbors,
                coef_self=args.coef_self, coef_neighbors=args.coef_neighbors,
                self_runs=self_runs, verbose=args.verbose
            )
            run_total += written
            run_correct += correct

            # Per-user accuracy
            user_acc = (correct / written * 100.0) if written > 0 else 0.0
            print(f"User {userid} stats: written {written}, correct {correct}, accuracy {user_acc:.2f}%")

        # This run accuracy
        run_acc = (run_correct / run_total * 100.0) if run_total > 0 else 0.0
        print(f"\n== Run {run_num} summary ==")
        print(f"Samples written this run: {run_total}")
        print(f"Correct samples this run: {run_correct}")
        print(f"Accuracy this run: {run_acc:.2f}%")
        print(f"Results saved to: {output_csv}")

        # Add to global
        global_total += run_total
        global_correct += run_correct

    # Global accuracy
    global_acc = (global_correct / global_total * 100.0) if global_total > 0 else 0.0
    print("\n================ Global Summary ================")
    print(f"Total number of test samples: {global_total}")
    print(f"Total number of correct samples: {global_correct}")
    print(f"Global accuracy: {global_acc:.2f}%")
    print("===============================================")

